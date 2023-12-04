import sys # Modulo necesario para recoger datos de la linea de comandos
import openpyxl
from openpyxl import load_workbook  # Libreria para cargar el archivo
from openpyxl.styles import Font, NumberFormatDescriptor  # Libreria para cambiar la fuente
from openpyxl.styles import DEFAULT_FONT
###########################################
################## Ayuda ##################
###########################################
help = """
[Parametros]
-h: Muestra el manual

-T: Cambia el tamaño de las columnas a las apropiadas para el programa

-F: Cambia la fuente y el tamaño de la fuente a Arial 10

-P: Cambia las propiedades de las columnas de precio y codigo de barras

-M: Aplica mayusculas a todas las descripciones

-R: Borra los caracteres especiales de los numeros de referencia, dejando solo los numeros\n    y reemplaza las vocales con tilde, la Ñ, los espacios antes y despues de los caracteres especiales\n    en las descripciones\n

-E Cambia la extension del archivo a la que indiques
[Funcionamiento]
Ejemplo: python Tarifeitor.py ejemplo.xlsx -TPF -e xls
En este ejemplo se ejecutaria el cambio de tamaño de columnas(T), cambio de propiedades de las columnas de precio(P)\n""" + """y cambio de fuentes, ademas se exportaria en formato xls(F)\n
El primer parametro debe ser el libro con el que trabajaremos, tiene que estar ya ordenado en este orden:\n
Codigo | Indice | Codigo de Fabricante | Descripcion | Subfamilia | PV1 | PV2 | Precio Neto | Codigo de Barras\n
El segundo parametro puedes dejarlo en blanco y ejecutara automaticamente todas las funciones, en caso de que quieras\n""" + """ejecutar funciones concretas, debes especificarlas como en el ejemplo\n
Puedes añadir separado el parametro -E para especificar la extension de archivo a utilizar, por defecto se usara la extension original"""

###########################################
################ Funciones ################
###########################################

# columnSize()
# fontAndSize()
# replaceCharacters()
# columnProperty()
# upperCase()

def changeColumnSize(ws):  # Cambia el tamaño de las columnas
    print("Cambiando tamaño de las columnas...")
    columnSizes = {
        "A": 20.69, "B": 20.69, "C": 20.69, "D": 60.69,
        "E": 5.69, "F": 6.69, "G": 12.69, "H": 12.69,
        "I": 12.69, "J": 13.69
    }

    for col, width in columnSizes.items():
        ws.column_dimensions[col].width = width


def changeFontAndSize(ws): # Cambia la fuente y el tamaño de fuente de todas las columnas
    print("Cambiado fuente y tamaño de letra...")
    font = Font(name='Arial', size=10)
    for col in ws.iter_cols(min_col=1, max_col=10):
        for cell in col:
            cell.font = font

def changeColumnProperty(ws): # Cambia el tipo de columna
    print("Cambiando propiedades de las columnas")
    precios = [7, 8, 9]
    for col in precios: # Recorre las columnas de los precios
        for row in range(1, ws.max_row + 1):
            x = ws.cell(column=col, row=row).value
            if x is not None:
                ws.cell(column=col, row=row).value = float(x) # Reconvierte el numero a float y lo guarda en la celda
                xfinal = str(x)[-2:] # Extrae el final del numero y lo convierte en string para poder concatenarlo e iterarlo
                xinicio = str(x)[0:-2] # Extrae el principio del numero
                if "." in xinicio: #Si el numero ya tiene punto lo junta sin mas, si no lo tiene se lo añade
                    x = xinicio + xfinal
                else:
                    x = xinicio + "." + xfinal # Junta los numeros añadiendole el punto antes de los dos ultimos digitos
                ws.cell(column=col, row=row).number_format = "0.00" # Aplica el formato de celda numero y le pone dos decimales
        for row in range(1, ws.max_row + 1):
            ws.cell(column=10, row=row).number_format = "0" #Aplica el formato a la columna del numero de barras

def convertToUpperCase(ws): # Pone en Mayusculas toda la descripcion
    print("Cambiando letras a mayusculas...")
    for row in range(1, ws.max_row + 1):
        x = ws.cell(column=4, row=row).value
        if isinstance(x, str):
            ws.cell(column=4, row=row).value = x.upper()

def replaceCharacters(ws): # Reemplaza los caracteres indeseados
    print("Filtrando caracteres...")
    
    reemplazos = { # Diccionario con cada caracter a buscar y su reemplazo
        " / ": "/", " /": "/", "/ ": "/", " , ": ",", " ,": ",", ", ": ",", " . ": ".", " .": ".", ". ": ".",
        "á": "A", "é": "E", "í": "I", "ó": "O", "ú": "U", "ü": "U", "ñ": "N", "Ø": "D."
    }

    for row in range(1, ws.max_row + 1):
        x = ws.cell(column=4, row=row).value
        if x is not None: # Comprueba que la celda de excel no este vacia para evitar errores
            x = str(x)
            for buscar, reemplazo in reemplazos.items():
                x = x.replace(buscar, reemplazo)
            ws.cell(column=4, row=row).value = x

def saveDocument():
    ruta = str(sys.argv[1])
    barra = ruta.rfind("\\")
    if barra == -1:
        nCompleto = ruta[barra:]
    else:
        nCompleto = ruta
    dot = nCompleto.rfind(".")
    name = nCompleto[:dot]
    extension = nCompleto[dot+1:]
    finalFile = name + "Final." + extension
    return(finalFile)

##########################################
############### Parametros ###############
##########################################

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print (help + "\n\nDebes proporcionar el nombre del archivo Excel como argumento.")
        sys.exit(1)

    if sys.argv[1] == "-h":
        print(help)
        sys.exit(0)
    
    filename = sys.argv[1]

    try:
        wb = load_workbook(filename) # Define el libro de excel a utilizar como el especificado en la linea de comandos
        ws = wb.active

        command = None
        contador = 0
        if len(sys.argv) > 2:
            command = sys.argv[2].upper()

            # Procesando parametros #

            if "T" in command: 
                    changeColumnSize(ws)
                    contador +=1

            if "F" in command: 
                changeFontAndSize(ws)
                contador +=1

            if "P" in command: 
                    changeColumnProperty(ws)
                    contador +=1

            if "R" in command: 
                replaceCharacters(ws)
                contador +=1

            if "M" in command: 
                convertToUpperCase(ws)
                contador +=1

        if contador == 0:
            changeColumnSize(ws)
            changeFontAndSize(ws)
            replaceCharacters(ws)
            changeColumnProperty(ws)
            convertToUpperCase(ws)

        finalFile = saveDocument()
        wb.save(finalFile)
        print("Operacion exitosa. El archivo se ha guardado como:" + finalFile)
        sys.exit(0)
    except Exception as e:
        print("Error:", str(e))
        sys.exit(1)
